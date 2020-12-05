#first installed the module writing the code in the terminal -> pip install openpyxl

from openpyxl import Workbook
Workbook = Workbook()
Sheet = Workbook.active

Sheet["A1"] = "Date"
Sheet["B1"] = "Calibration"
Sheet["C1"] = "P1"

Workbook.save(filename="MTR_SWARM.xlsx")


#for the second sheet
#from openpyxl import load_workbook
#Workbook = load_workbook(filename="MTR_SWARM.xlsx")
#Workbook.sheetnames
#['Sheet 2']
#sheet = workbook.active
#Sheet
#<Worksheet "Sheet 2">
#Sheet.title
#'Sheet 2'

# Creating the second sheet or any sheet
Workbook.sheetnames
Operations_sheet = Workbook.create_sheet("11/29/2020")

#Opening a specific sheet
Sheet = Workbook.active